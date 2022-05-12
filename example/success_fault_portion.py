import openpyxl

from volleyball.stat.StatMaker import *
from openpyxl import *

stat_file = StatFile("C:\\Users\\user\\Documents\\배구", ["여자부"], [2021], list(range(1, 300)))
stat_file.fill_games()
stat_file.fill_sets()
stat_file.fill_messages()
stat_file.fill_rallies()
stat_file.fill_chances()

success_fault_portion_EffCalc = EffCalc({})


class GeneralAttackPlayer(Player):
    def __init__(self
                 , name: str
                 , back_num: int
                 , season: int
                 , team: str
                 , setter: str
                 , order: int
                 , attacker_position: int
                 , attack_type):
        super().__init__(name, back_num, season, team)
        self.setter = setter
        self.order = order
        self.attacker_position = attacker_position
        self.attack_type = take_key_by_value(action_to_class, attack_type)


def processor_in_set_messages_by_messages(messages: Messages
                                          , chances: Chances
                                          , rallies: Rallies
                                          , game: Game
                                          , set_in_game: Set
                                          , side: int):
    for message in messages:
        if isinstance(message, AttackMessage):
            chance = chances.select_chance_with_message_index(message.index)
            rally = rallies.select_rally_inner_message_index(message.index)
            if chance.message_index[1] - chance.message_index[0] == 0:
                setter = ""
                order = 1
            elif chance.message_index[1] - chance.message_index[0] == 1:
                setter = messages[message.index - 1].name
                order = 2
            else:
                setter = messages[message.index - 1].name
                order = 3

            try:
                success_fault_portion_EffCalc.append_effect_calc(Messages([message])
                                                                 , game
                                                                 , "공격종합"
                                                                 , GeneralAttackPlayer
                                                                 , setter
                                                                 , order
                                                                 , rally.rotation[side].index(message.name)
                                                                 , type(message))
            except ValueError:
                success_fault_portion_EffCalc.append_effect_calc(Messages([message])
                                                                 , game
                                                                 , "공격종합"
                                                                 , GeneralAttackPlayer
                                                                 , setter
                                                                 , order
                                                                 , 7
                                                                 , type(message))


stat_file.process_in_set_messages_by_messages(processor_in_set_messages_by_messages)

wb = openpyxl.Workbook()
sheet = wb.create_sheet(title="전체")

bigger_than_fifty = EffCalc(filter(lambda it: it[1][0] > 50, success_fault_portion_EffCalc.items()))

for i, stat in enumerate(list(map(lambda it: make_percentage(it[1][1], it[1][2]),  bigger_than_fifty.items()))):
    sheet.cell(row=1, column=i+1).value = stat

print(list(map(lambda it: make_percentage(it[1][1], it[1][2]),  bigger_than_fifty.items())))

print("---------------------------------")

sheet = wb.create_sheet(title="세터 별")

setters = list(dict.fromkeys(list(map(lambda it: it[0].setter, success_fault_portion_EffCalc.items()))))

setter_index = 0
for one_setter in setters:
    attack_after_setter = EffCalc(filter(lambda it: it[0].setter == one_setter and it[1][0] > 50
                                         , success_fault_portion_EffCalc.items()))
    if len(list(map(lambda it: make_percentage(it[1][1], it[1][2]), attack_after_setter.items()))) > 0:
        print(one_setter, list(map(lambda it: make_percentage(it[1][1], it[1][2]), attack_after_setter.items())))
    if len(list(map(lambda it: make_percentage(it[1][1], it[1][2]), attack_after_setter.items()))) > 0:
        sheet.cell(row=setter_index*2+1, column=1).value = one_setter
        for v, stat in enumerate(list(map(lambda it: make_percentage(it[1][1], it[1][2]), attack_after_setter.items()))):
            sheet.cell(row=setter_index*2+1, column=v+2).value = stat
            sheet.cell(row=setter_index*2+2, column=1).value = "평균"
            sheet.cell(row=setter_index*2+2, column=3).value = "편차"
        setter_index += 1

print("-------------------------------")

sheet = wb.create_sheet(title="공격 종류 별")

for i, attack_type in enumerate(attack_types[:3]):
    attack_type_attack = EffCalc(filter(lambda it: it[0].attack_type == attack_type and it[1][0] > 50
                                        , success_fault_portion_EffCalc.items()))
    print(attack_type, list(map(lambda it: make_percentage(it[1][1], it[1][2]), attack_type_attack.items())))
    sheet.cell(row=i*2+1, column=1).value = attack_type
    for v, stat in enumerate(list(map(lambda it: make_percentage(it[1][1], it[1][2]), attack_type_attack.items()))):
        sheet.cell(row=i*2+1, column=v+2).value = stat
        sheet.cell(row=i * 2 + 2, column=1).value = "평균"
        sheet.cell(row=i * 2 + 2, column=3).value = "편차"


print("-------------------------------")
sheet = wb.create_sheet(title="전후위")
front = [1, 2, 3]
back = [0, 4, 5]

front_attack = EffCalc(filter(lambda it: it[0].attacker_position in front and it[1][0] > 50
                              , success_fault_portion_EffCalc.items()))
print("전위", list(map(lambda it: make_percentage(it[1][1], it[1][2]), front_attack.items())))
sheet.cell(row=1, column=1).value = "전위"
for i, stat in enumerate(list(map(lambda it: make_percentage(it[1][1], it[1][2]), front_attack.items()))):
    sheet.cell(row=1, column=i+2).value = stat
sheet.cell(row=2, column=1).value = "평균"
sheet.cell(row=2, column=3).value = "편차"

back_attack = EffCalc(filter(lambda it: it[0].attacker_position in back and it[1][0] > 50
                             , success_fault_portion_EffCalc.items()))
print("후위", list(map(lambda it: make_percentage(it[1][1], it[1][2]), back_attack.items())))
sheet.cell(row=3, column=1).value = "후위"
for i, stat in enumerate(list(map(lambda it: make_percentage(it[1][1], it[1][2]), back_attack.items()))):
    sheet.cell(row=3, column=i+2).value = stat
sheet.cell(row=4, column=1).value = "평균"
sheet.cell(row=4, column=3).value = "편차"

wb.save("공격.xlsx")
wb.close()
